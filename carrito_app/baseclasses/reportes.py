from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import sqlite3
from datetime import date

import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import formatdate
from email import encoders


from jnius import autoclass
Environment = autoclass('android.os.Environment')
path = Environment.getExternalStorageDirectory().getAbsolutePath()


class SuccessPopup(Popup):
    pass


class ReportesWid(BoxLayout):
    report = ''

    def __init__(self, mainwid, **kwargs):
        super(ReportesWid, self).__init__()
        self.mainwid = mainwid

    def checkbox_click(self, instance, value, report):
        try:
            if value == True:
                ReportesWid.report = report
            else:
                ReportesWid.report = ''

        except Exception as e:
            print(e)

    def get_report(self, mes, año):
        con = sqlite3.connect('database/carrito.db')
        cursor = con.cursor()

        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        if mes == 'Todos':
            a = (ReportesWid.report, año)
            print(a)

            s1 = f"""
                SELECT * FROM {ReportesWid.report} 
                WHERE fecha BETWEEN '{año}-01-01' AND '{año}-12-31'
            """
            cursor.execute(s1)

            result = cursor.fetchall()
            print(result)

        else:
            mes_int = meses.index(mes) + 1
            a = (ReportesWid.report, mes_int, año)
            print(a)

            s2 = f"""
                SELECT * FROM {ReportesWid.report}
                WHERE fecha BETWEEN '{año}-{mes_int}-01' AND '{año}-{mes_int}-31'
            """
            cursor.execute(s2)

            result = cursor.fetchall()
            print(result)

        if ReportesWid.report == 'insumos':
            self.export_to_excel_insumos(result)

        if ReportesWid.report == 'ventas':
            self.export_to_excel_ventas(result)

        con.close()

    def export_to_excel_insumos(self, result):
        fecha_reporte = date.today()
        filename = f'{self.ids.ti_asunto.text}_{fecha_reporte}.xlsx'

        wb = Workbook()

        hoja = wb.active
        hoja.title = f"Reporte de {ReportesWid.report}_{fecha_reporte}"

        asunto = f"Reporte de {ReportesWid.report}_{fecha_reporte}"

        hoja['A1'] = 'data_id'
        hoja['B1'] = 'producto'
        hoja['C1'] = 'cantidad'
        hoja['D1'] = 'unidad'
        hoja['E1'] = 'categoria'
        hoja['F1'] = 'costo'
        hoja['G1'] = 'fecha'

        hoja['A1'].font = Font(bold=True)
        hoja['B1'].font = Font(bold=True)
        hoja['C1'].font = Font(bold=True)
        hoja['D1'].font = Font(bold=True)
        hoja['E1'].font = Font(bold=True)
        hoja['F1'].font = Font(bold=True)
        hoja['G1'].font = Font(bold=True)

        fila = 2  # Fila donde empezamos
        col_id = 1
        col_producto = 2
        col_cantidad = 3
        col_unidad = 4
        col_categoria = 5
        col_costo = 6
        col_fecha = 7

        for data_id, producto, cantidad, unidad, categoria, costo, fecha in result:
            hoja.cell(column=col_id, row=fila, value=data_id)
            hoja.cell(column=col_producto, row=fila, value=producto)
            hoja.cell(column=col_cantidad, row=fila, value=cantidad)
            hoja.cell(column=col_unidad, row=fila, value=unidad)
            hoja.cell(column=col_categoria, row=fila, value=categoria)
            hoja.cell(column=col_costo, row=fila, value=costo)
            hoja.cell(column=col_fecha, row=fila, value=fecha)
            fila += 1

        ruta = f'{path}/Download/{filename}'
        wb.save(filename=ruta)
        self.mainwid.goto_start()

        message = self.mainwid.Popup.ids.message
        self.mainwid.Popup.open()
        self.mainwid.Popup.title = 'Generación de Reporte'
        txt1 = '**** Reporte de Insumos ****'
        txt2 = 'Se ha generado exitosamente'

        message.text = f'{txt1}\n{txt2}'

        self.ids.ti_asunto.text = ''

        print(f'Sending data to: {ruta}')

        self.sendmail(asunto, ruta, filename)

    def export_to_excel_ventas(self, result):
        fecha_reporte = date.today()
        filename = f'{self.ids.ti_asunto.text}_{fecha_reporte}.xlsx'

        wb = Workbook()

        hoja = wb.active
        hoja.title = f"Reporte de {ReportesWid.report}_{fecha_reporte}"

        asunto = f"Reporte de {ReportesWid.report}_{fecha_reporte}"

        hoja['A1'] = 'ID'
        hoja['B1'] = 'Pedido'
        hoja['C1'] = 'Total'
        hoja['D1'] = 'Forma de Pago'
        hoja['E1'] = 'Fecha'

        hoja['A1'].font = Font(bold=True)
        hoja['B1'].font = Font(bold=True)
        hoja['C1'].font = Font(bold=True)
        hoja['D1'].font = Font(bold=True)
        hoja['E1'].font = Font(bold=True)

        fila = 2  # Fila donde empezamos
        col_id = 1
        col_pedido = 2
        col_total = 3
        col_fpago = 4
        col_fecha = 5

        for data_id, pedido, total, fpago, fecha in result:
            hoja.cell(column=col_id, row=fila, value=data_id)
            hoja.cell(column=col_pedido, row=fila, value=pedido)
            hoja.cell(column=col_total, row=fila, value=total)
            hoja.cell(column=col_fpago, row=fila, value=fpago)
            hoja.cell(column=col_fecha, row=fila, value=fecha)
            fila += 1

        ruta = f'{path}/Download/{filename}'
        wb.save(filename=ruta)
        self.mainwid.goto_start()

        message = self.mainwid.Popup.ids.message
        self.mainwid.Popup.open()
        self.mainwid.Popup.title = 'Generación de Reporte'
        txt1 = '**** Reporte de Ventas **** '
        txt2 = 'Se ha generado exitosamente'

        message.text = f'{txt1}\n{txt2}'
        self.ids.ti_asunto.text = ''

        print(f'Sending data to: {ruta}')

        self.sendmail(asunto, ruta, filename)

    def sendmail(self, asunto, ruta, filename):
        username = 'tucorreodedestino@gmail.com'
        password = 'tucontraseña'
        send_from = 'tucorreodeorigen@gmail.com'
        send_to = 'tucorreodedestino@gmail.com'
        Cc = ''

        msg = MIMEMultipart()
        msg['From'] = send_from
        msg['To'] = send_to
        msg['Cc'] = Cc
        msg['Date'] = formatdate(localtime=True)
        msg['Subject'] = asunto

        fp = open(ruta, 'rb')
        part = MIMEBase('application', 'vnd.ms-Excel')
        part.set_payload(fp.read())
        fp.close()

        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=filename)
        msg.attach(part)

        smtp = smtplib.SMTP('smtp.gmail.com', 587)
        smtp.starttls()
        smtp.login(username, password)
        smtp.sendmail(send_from, send_to.split(',') +
                      msg['Cc'].split(','), msg.as_string())
        smtp.quit()

        print('Mensaje enviado...')
